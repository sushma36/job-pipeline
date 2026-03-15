"""
=============================================================
  excel_exporter.py — Exports matched jobs to Excel
=============================================================
"""

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime


# ─── COLORS ──────────────────────────────────────────────────
C_NAVY      = "13315C"
C_HEADER    = "1E3A5F"
C_WHITE     = "FFFFFF"
C_ALT       = "F2F7FC"
C_GREEN_BG  = "D6EFCD"
C_AMBER_BG  = "FFF3CD"
C_RED_BG    = "FCE4D6"
C_INFO_BG   = "EBF3FB"
C_BORDER    = "B0C4DE"

COLUMNS = [
    ("Job Title",       30),
    ("Company",         22),
    ("Platform",        18),
    ("Location",        22),
    ("Work Mode",       14),
    ("Posting Date",    16),
    ("Match Score",     13),
    ("Matched Skills",  42),
    ("Missing Skills",  32),
    ("Recommendation",  16),
    ("Job URL",         45),
]


def _border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def export_to_excel(jobs: list, output_path: str) -> str:
    wb = Workbook()
    _build_jobs_sheet(wb, jobs)
    _build_guide_sheet(wb, jobs)
    wb.save(output_path)
    print(f"✅ Exported {len(jobs)} jobs → {output_path}")
    return output_path


def _build_jobs_sheet(wb, jobs):
    ws = wb.active
    ws.title = "🔍 Job Matches"

    # ── Title banner ──────────────────────────────────────────
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value     = "🚀  DATA ENGINEER JOB MATCHES  •  Sushma Dasari  •  Last 48 Hours"
    c.font      = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Metadata ─────────────────────────────────────────────
    ws.merge_cells("A2:K2")
    m = ws["A2"]
    m.value     = (f"Generated: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}  •  "
                   f"{len(jobs)} jobs matched  •  Sorted: Newest first, then by score")
    m.font      = Font(name="Arial", italic=True, size=10, color="555555")
    m.fill      = PatternFill("solid", fgColor=C_INFO_BG)
    m.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Stats row ─────────────────────────────────────────────
    ws.row_dimensions[3].height = 8   # spacer
    ws.row_dimensions[4].height = 8   # spacer

    # ── Column headers ────────────────────────────────────────
    HDR_ROW = 5
    for ci, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=name)
        cell.font      = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HDR_ROW].height = 28

    # ── Data rows ─────────────────────────────────────────────
    for ri, job in enumerate(jobs):
        row  = HDR_ROW + 1 + ri
        alt  = ri % 2 == 1
        rec  = job.get("recommendation", "")
        bg   = (C_GREEN_BG if "Apply" in rec
                else C_AMBER_BG if "Consider" in rec
                else C_ALT if alt else C_WHITE)
        score = job.get("match_score", 0)
        score_color = ("375623" if score >= 80
                       else "7F6000" if score >= 65
                       else "595959")

        values = [
            job.get("job_title", ""),
            job.get("company_name", ""),
            job.get("platform_name", ""),
            job.get("location", ""),
            job.get("remote_or_hybrid", ""),
            job.get("posting_date", "")[:10] if job.get("posting_date") else "",
            score,
            job.get("matched_skills", ""),
            job.get("missing_skills", ""),
            rec,
            job.get("job_url", ""),
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font      = Font(name="Arial", size=9)

            if ci != 7:  # not score column
                cell.fill = PatternFill("solid", fgColor=bg)

            # Score column
            if ci == 7:
                cell.font      = Font(name="Arial", bold=True, size=11,
                                      color=score_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill      = PatternFill("solid", fgColor=bg)

            # Recommendation column
            if ci == 10:
                rc = ("375623" if "Apply" in rec
                      else "7F6000" if "Consider" in rec
                      else "8B0000")
                cell.font      = Font(name="Arial", bold=True, size=10, color=rc)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # URL column — hyperlink
            if ci == 11 and val:
                cell.hyperlink = val
                cell.font      = Font(name="Arial", size=9,
                                      color="0563C1", underline="single")
                cell.value     = "🔗 Open"

            # Date column
            if ci == 6:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row].height = 52

    # ── Freeze & filter ───────────────────────────────────────
    ws.freeze_panes = f"A{HDR_ROW + 1}"
    last = HDR_ROW + len(jobs)
    ws.auto_filter.ref = f"A{HDR_ROW}:K{last}"

    # ── Color scale on score column ───────────────────────────
    if jobs:
        ws.conditional_formatting.add(
            f"G{HDR_ROW+1}:G{last}",
            ColorScaleRule(
                start_type="num", start_value=40,  start_color="FCE4D6",
                mid_type="num",   mid_value=65,    mid_color="FFEB9C",
                end_type="num",   end_value=100,   end_color="C6EFCE",
            )
        )


def _build_guide_sheet(wb, jobs):
    ws = wb.create_sheet("📋 Scoring Guide")

    def hdr(row, col, text, span=4):
        ws.merge_cells(f"{get_column_letter(col)}{row}:"
                       f"{get_column_letter(col+span-1)}{row}")
        c = ws.cell(row=row, column=col, value=text)
        c.font      = Font(name="Arial", bold=True, size=11, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
        ws.row_dimensions[row].height = 22

    def row(r, col, label, value, bg=C_WHITE):
        lc = ws.cell(row=r, column=col, value=label)
        lc.font   = Font(name="Arial", bold=True, size=10)
        lc.fill   = PatternFill("solid", fgColor=bg)
        lc.border = _border()
        vc = ws.cell(row=r, column=col+1, value=value)
        vc.font      = Font(name="Arial", size=10)
        vc.fill      = PatternFill("solid", fgColor=bg)
        vc.border    = _border()
        vc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 18

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = "Scoring Guide & Run Statistics"
    t.font      = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    t.fill      = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Weights
    r = 3
    hdr(r, 1, "📊 Match Score Weights")
    weights = [
        ("Skills Overlap",       "40%", "Resume skills found in JD"),
        ("Tech Stack Match",     "25%", "Python, Spark, Airflow, Snowflake, AWS..."),
        ("Experience Level",     "15%", "Junior/Mid = high; Senior/Management = lower"),
        ("Domain Relevance",     "10%", "Healthcare, enterprise, analytics, cloud"),
        ("Education Req.",       "10%", "MS CS = full score for BS/MS roles"),
    ]
    alt = ["F2F7FC", "FFFFFF"]
    for i, (name, pct, desc) in enumerate(weights):
        r += 1
        bg = alt[i % 2]
        ws.cell(row=r, column=1, value=name).font  = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=1).fill              = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=1).border            = _border()
        ws.cell(row=r, column=2, value=pct).font   = Font(name="Arial", bold=True,
                                                           size=10, color="1E3A5F")
        ws.cell(row=r, column=2).alignment         = Alignment(horizontal="center")
        ws.cell(row=r, column=2).fill              = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=2).border            = _border()
        ws.cell(row=r, column=3, value=desc).font  = Font(name="Arial", size=10)
        ws.cell(row=r, column=3).fill              = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=3).border            = _border()
        ws.row_dimensions[r].height = 18

    # Platforms
    r += 2
    hdr(r, 1, "🌐 Data Sources")
    sources = [
        ("Greenhouse ATS", "Direct JSON API",  "50+ companies — Stripe, Databricks, Figma..."),
        ("Lever ATS",      "Direct JSON API",  "Netflix, Confluent, ClickHouse..."),
        ("Remotive",       "Direct JSON API",  "Remote tech jobs"),
        ("RemoteOK",       "Direct JSON API",  "Remote-first, multiple tags"),
        ("WeWorkRemotely", "Direct RSS",        "Curated remote jobs, 3 feeds"),
        ("Jobicy",         "Direct JSON API",  "Remote jobs, multiple tags"),
        ("SmartRecruiters","Direct REST API",  "Snowflake, HubSpot, Okta..."),
        ("Indeed",         "Apify Actor",      "Largest job board, 24hr filter"),
        ("MyVisaJobs",     "Apify Actor",      "H1B sponsoring employers"),
    ]
    for i, (name, stype, desc) in enumerate(sources):
        r += 1
        bg = alt[i % 2]
        ws.cell(row=r, column=1, value=name).font  = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=1).fill              = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=1).border            = _border()
        ws.cell(row=r, column=2, value=stype).font = Font(name="Arial", size=10,
                                                          color="1E3A5F")
        ws.cell(row=r, column=2).fill              = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=2).border            = _border()
        ws.cell(row=r, column=3, value=desc).font  = Font(name="Arial", size=10)
        ws.cell(row=r, column=3).fill              = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=3).border            = _border()
        ws.row_dimensions[r].height = 18

    # Stats
    r += 2
    hdr(r, 1, "📈 Run Statistics", span=2)
    from collections import Counter
    platform_counts = Counter(j.get("platform_name", "") for j in jobs)
    apply   = sum(1 for j in jobs if "Apply" in j.get("recommendation", ""))
    consider = sum(1 for j in jobs if "Consider" in j.get("recommendation", ""))
    avg     = round(sum(j.get("match_score", 0) for j in jobs) / max(len(jobs), 1))
    stats   = [
        ("Total Jobs",             len(jobs)),
        ("✅ Apply",               apply),
        ("⚠️ Consider",            consider),
        ("Average Match Score",    f"{avg}/100"),
        ("Generated",              datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (label, val) in enumerate(stats):
        r += 1
        row(r, 1, label, val, bg=alt[i % 2])

    r += 2
    hdr(r, 1, "📊 Jobs by Platform", span=2)
    for i, (platform, count) in enumerate(
            sorted(platform_counts.items(), key=lambda x: -x[1])):
        r += 1
        row(r, 1, platform, count, bg=alt[i % 2])

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
